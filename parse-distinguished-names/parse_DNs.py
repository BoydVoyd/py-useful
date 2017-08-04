import sys
import os.path
import ldap.dn
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string

if (len(sys.argv) >1):
    ws_file = sys.argv[1]
else:
    ws_file = './EarlyWorkstations.xlsx'

if os.path.isfile(ws_file):
    pass
else:
    print "File Not Found"
    sys.exit()

out_file = os.path.dirname(ws_file) + "/" +os.path.splitext(os.path.basename(ws_file))[0] + "_parsed.xlsx"

wb = openpyxl.load_workbook(ws_file)
sheet = wb.get_sheet_by_name(wb.get_sheet_names()[0])
for i1 in range(2, sheet.max_row +1):
    ou = sheet['B' + str(i1)].value
    ou_elements = ldap.dn.str2dn(ou)
    ou_elements.reverse()
    for i2 in range(len(ou_elements)):
        e = (ou_elements)[i2][0][1]
        sheet[get_column_letter(3+i2) + str(i1)] = e
wb.save(filename=out_file)